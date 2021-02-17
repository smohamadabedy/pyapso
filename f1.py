##==============================================================
##==============================================================
##===================Adaptive PSO Algorithm=====================
##=========create in 2021/feb by m.h.o.abedi@gmail.com==========
##========================free_to_use===========================
##this algorithm is base on wieght coefficent correction every
##iteration save on seprated excel workbook
##initilization need pop size / numb of iteration and coeff value
import time
from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()  
sheet = wb.active
sheet.title = "0"
import numpy as np
from fit_func import *
myfunc("hello welcome to APSO Optimization")    
num = input("number of iteration : \n\r")
#initiall_p1
#Generate population
#a = np.vstack((a,np.array([1,2,3])))
start = time.time()
pop_size=100
iteration=20
max_wg = .8
min_wg = .6
C1 = .8;
C2 = 2;
space_num = np.array([10,15])
N_dim = list(space_num.shape)
particle_group = {}
for i in range(0,pop_size):
    pos =[];
    vel =[];
    
    for j in range(0,N_dim[0]):
        pos.append(np.random.rand()*space_num[j])
        vel.append(np.random.uniform(-1,1))
        sheet.cell(row=i+1, column=j+1).value = pos[j]
        sheet.cell(row=i+1, column=j+1+N_dim[0]+1).value = vel[j]
        
    particle_group[i] = {
        "pos"  : np.array(pos),
        "vel"  : np.array(vel),
        "fval" : np.array([func(pos)]),
        "pbest" : np.append(pos,[func(pos)]),
        "w_inertia" : 0
        };
 
    sheet.cell(row=1+i, column=N_dim[0]*2+6).value = particle_group[i]["fval"][0]
    
   
sheet.cell(row=i+2, column=N_dim[0]*2+6).fill =  PatternFill(start_color="49E64F", end_color="49E64F", fill_type = "solid")
sheet.cell(row=i+2, column=N_dim[0]*2+6).value  = f"=MIN(OFFSET(A1,0,{N_dim[0]*2+6-1}):OFFSET(A1,{i},{N_dim[0]*2+6-1}))"; 

bg = find_bestg(particle_group)
i = 0
F_bar = 0;
for m in particle_group:
    F_bar += particle_group[m]["fval"][0]; 
    i = i+1
Fbar = F_bar/i;
for m in particle_group:
    F_bar += particle_group[m]["fval"][0]; 
    i = i+1
print("********")
for m in particle_group:
    if(particle_group[m]["fval"][0] <= F_bar):
        particle_group[m]["w_inertia"] = min_wg + (((particle_group[m]["fval"][0] - bg["gbest"])*(max_wg-min_wg))/(Fbar-bg["gbest"]))
    else:
        particle_group[m]["w_inertia"] = max_wg
    sheet.cell(row=1+m, column=N_dim[0]*2+3).value =  particle_group[m]["w_inertia"]

for it in range(1,int(num)):
    sh = wb.create_sheet(title=str(it))
    for part in particle_group:
        pos =[];
        vel =[];
        
        for j in range(0,N_dim[0]):
            V_part = (particle_group[part]["w_inertia"]*particle_group[part]["vel"][j]) +(C1*(particle_group[part]["pbest"][j]-particle_group[part]["pos"][j])) +(C2*(bg["pos"][j]-particle_group[part]["pos"][j]))
            pos.append(V_part + particle_group[part]["pos"][j] )
            vel.append(V_part)
            sh.cell(row=part+1, column=j+1).value = pos[j]
            sh.cell(row=part+1, column=j+1+N_dim[0]+1).value = vel[j]
            
        fvalue = func(pos)
        
        if (fvalue < particle_group[part]["pbest"][N_dim][0]):
            particle_group[part]["pbest"] = np.append(pos,[func(pos)])
      
        particle_group[part]["pos"] =np.array(pos)
        particle_group[part]["vel"] =np.array(vel)
        particle_group[part]["fval"] =np.array([fvalue])
        sh.cell(row=part+1, column=N_dim[0]*2+6).value = particle_group[part]["fval"][0]

    sh.cell(row=part+2, column=N_dim[0]*2+6).fill =  PatternFill(start_color="49E64F", end_color="49E64F", fill_type = "solid")
    sh.cell(row=part+2, column=N_dim[0]*2+6).value  = f"=MIN(OFFSET(A1,0,{N_dim[0]*2+6-1}):OFFSET(A1,{part},{N_dim[0]*2+6-1}))";
    bg_f = find_bestg(particle_group)

    if(bg_f['gbest'] < bg['gbest']):
        bg = bg_f
    else:
        pass
    i = 0
    F_bar = 0;
    for m in particle_group:
        F_bar += particle_group[m]["fval"][0]; 
        i = i+1
    Fbar = F_bar/i;
    for m in particle_group:
        F_bar += particle_group[m]["fval"][0]; 
        i = i+1
    for m in particle_group:
        if(particle_group[m]["fval"][0] <= F_bar):
            particle_group[m]["w_inertia"] = min_wg + (((particle_group[m]["fval"][0] - bg["gbest"])*(max_wg-min_wg))/(Fbar-bg["gbest"]))
        else:
            particle_group[m]["w_inertia"] = max_wg
        sh.cell(row=1+m, column=N_dim[0]*2+3).value =  particle_group[m]["w_inertia"]    
print("*"*10)
print("-"*10)
print("~"*10)
print("end")
stop = time.time()
print('Time: ', stop - start)
print(it)
wb.save("sample_file1.xlsx")  
input()





